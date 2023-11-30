import datetime
from pprint import pprint

import openpyxl
from aiogram import types
from aiogram.fsm.context import FSMContext
from openpyxl.styles import Alignment, Side, Border, Font, PatternFill

from sqlalchemy import CursorResult, select
from sqlalchemy.orm import sessionmaker, selectinload

from bot.commands.login_handlers import start
from bot.db.models import Order, Pharmacy


async def create_act(message: types.Message, state: FSMContext, session_maker: sessionmaker) -> None:
    data = await state.get_data()
    async with session_maker() as session:
        query: CursorResult = await session.execute(
            select(Order).join(Pharmacy, Order.pharmacy_id == Pharmacy.id).filter(Pharmacy.name == data['pharmacy'],
                                                                                  Pharmacy.district == data['district']) \
                .options(selectinload(Order.medicine), selectinload(Order.pharmacy)))

        orders = query.scalars().all()
        order: Order = orders[0]
        district = order.pharmacy.district
        kontra_agent = order.pharmacy.name

        ret_data = {
            'duty': 0,
            'sum_leftovers': 0,
            'sum_difference': 0,
            'district': district,
            'kontra_agent': kontra_agent,
            'table': []

        }
        # {
        #     'goods': [],
        #     'quantity': [],
        #     'difference': [],
        #     'leftover': []
        # }
        for order in orders:
            good = order.medicine.name

            json_good = {}

            json_good['good'] = good
            quantity = order.quantity
            json_good['quantity'] = quantity
            leftover = order.leftover
            json_good['leftover'] = leftover
            differance = quantity - leftover
            json_good['difference'] = differance
            json_good['price'] = order.medicine.price

            ret_data['table'].append(json_good)

            duty = quantity * order.medicine.price
            json_good['sum'] = duty
            ret_data['duty'] += duty

            leftovers = leftover * order.medicine.price
            json_good['sum_leftovers'] = leftovers
            ret_data['sum_leftovers'] += leftovers

            sum_difference = duty - leftovers
            ret_data['sum_difference'] += sum_difference
        pprint(ret_data)

        workbook = openpyxl.load_workbook('shablon.xlsx')
        sheet = workbook.active
        thick = Side(border_style="thin", color="000000")

        sheet[f'D5'].value = ret_data['district']
        sheet[f'D6'].value = ret_data['kontra_agent']
        sheet[f'D4'].value = ''

        for i, good in enumerate(ret_data['table'], 1):
            number_of_row = 8 + i
            if good['leftover'] == 0:
                fill = PatternFill(fill_type='lightUp',
                                   start_color='FAA0A0',
                                   end_color='FAA0A0')
                for g in sheet[f'B{number_of_row}:F{number_of_row}'][0]:
                    g.fill = fill
                pass

            sheet[f'B{number_of_row}'].value = i
            sheet[f'B{number_of_row}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'B{number_of_row}'].font = Font(size=11, color='000000')
            sheet[f'B{number_of_row}'].border = Border(top=thick, left=thick, right=thick, bottom=thick)

            sheet[f'C{number_of_row}'].value = good['good']
            sheet[f'C{number_of_row}'].font = Font(size=11, color='000000')
            sheet[f'C{number_of_row}'].border = Border(top=thick, left=thick, right=thick, bottom=thick)

            sheet[f'D{number_of_row}'].value = good['quantity']
            sheet[f'D{number_of_row}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'D{number_of_row}'].font = Font(size=11, color='000000')
            sheet[f'D{number_of_row}'].border = Border(top=thick, left=thick, right=thick, bottom=thick)

            sheet[f'E{number_of_row}'].value = good['leftover']
            sheet[f'E{number_of_row}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'E{number_of_row}'].font = Font(size=11, color='000000')
            sheet[f'E{number_of_row}'].border = Border(top=thick, left=thick, right=thick, bottom=thick)

            sheet[f'F{number_of_row}'].value = good['difference']
            sheet[f'F{number_of_row}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'F{number_of_row}'].font = Font(size=11, color='000000')
            sheet[f'F{number_of_row}'].border = Border(top=thick, left=thick, right=thick, bottom=thick)

            sheet[f'H{number_of_row}'].value = good['price']
            sheet[f'H{number_of_row}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'H{number_of_row}'].font = Font(size=11, color='000000')
            sheet[f'H{number_of_row}'].border = Border(top=thick, left=thick, right=thick, bottom=thick)

            sheet[f'I{number_of_row}'].value = good['sum']
            sheet[f'I{number_of_row}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'I{number_of_row}'].font = Font(size=11, color='000000')
            sheet[f'I{number_of_row}'].border = Border(top=thick, left=thick, right=thick, bottom=thick)

            sheet[f'J{number_of_row}'].value = good['sum_leftovers']
            sheet[f'J{number_of_row}'].alignment = Alignment(horizontal='center', vertical='center')
            sheet[f'J{number_of_row}'].font = Font(size=11, color='000000')
            sheet[f'J{number_of_row}'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
        number_of_row += 3
        sheet[f'B{number_of_row}'].value = f'Задолженность на {datetime.datetime.today().date()} (сумм)'
        sheet[f'B{number_of_row}'].font = Font(size=11, color='000000', bold=True)
        sheet[f'E{number_of_row}'].border = Border(left=thick, bottom=thick)
        sheet[f'B{number_of_row}'].border = Border(left=thick, right=thick, bottom=thick)
        sheet[f'C{number_of_row}'].border = Border(bottom=thick)
        sheet[f'D{number_of_row}'].border = Border(bottom=thick)
        sheet[f'F{number_of_row}'].border = Border(bottom=thick)
        sheet[f'E{number_of_row}'].font = Font(size=11, color='000000', bold=True)
        sheet[f'E{number_of_row}'].alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'E{number_of_row}'].value = ret_data['duty']
        number_of_row += 2
        sheet[f'B{number_of_row}'].value = f'Ост на {datetime.datetime.today().date()} (сумм)'
        sheet[f'B{number_of_row}'].font = Font(size=11, color='000000', bold=True)
        sheet[f'E{number_of_row}'].border = Border(left=thick, bottom=thick)
        sheet[f'B{number_of_row}'].border = Border(left=thick, right=thick, bottom=thick)
        sheet[f'C{number_of_row}'].border = Border(bottom=thick)
        sheet[f'D{number_of_row}'].border = Border(bottom=thick)
        sheet[f'F{number_of_row}'].border = Border(bottom=thick)
        sheet[f'E{number_of_row}'].alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'E{number_of_row}'].font = Font(size=11, color='000000', bold=True)
        sheet[f'E{number_of_row}'].value = ret_data['sum_leftovers']
        number_of_row += 2

        sheet[f'B{number_of_row}'].value = f'Разница на {datetime.datetime.today().date()} (сумм)'
        sheet[f'B{number_of_row}'].font = Font(size=11, color='000000', bold=True)
        sheet[f'E{number_of_row}'].border = Border(left=thick, bottom=thick)
        sheet[f'B{number_of_row}'].border = Border(left=thick, right=thick, bottom=thick)
        sheet[f'C{number_of_row}'].border = Border(bottom=thick)
        sheet[f'D{number_of_row}'].border = Border(bottom=thick)
        sheet[f'F{number_of_row}'].border = Border(bottom=thick)
        sheet[f'E{number_of_row}'].alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'E{number_of_row}'].font = Font(size=11, color='000000', bold=True)
        sheet[f'E{number_of_row}'].value = ret_data['sum_difference']
        workbook.save('test.xlsx')

        docs = types.FSInputFile('test.xlsx')

        await message.bot.send_document(message.from_user.id, document=docs)
        await state.clear()
        await start(message=message, session_maker=session_maker)
